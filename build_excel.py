"""
Génère le fichier Excel compte_resultat_prevoyance.xlsx
Compte de résultat formaté, table de tarification, et feuille de synthèse.
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from model_prevoyance import (
    build_mortality_table, simulate_portfolio, tarifier_portefeuille,
    simulate_sinistres, compte_resultat, compte_resultat_theorique,
    monte_carlo_sp, tarif_renouvellement
)

# ── Couleurs (hex sans #) ──────────────────────────────────────
NAVY_HEX    = "1F3864"
BLUE_HEX    = "2E75B6"
LIGHT_HEX   = "D6E4F0"
GREEN_HEX   = "70AD47"
ORANGE_HEX  = "C55A11"
GREY_HEX    = "F2F2F2"
WHITE_HEX   = "FFFFFF"

# ── Styles helpers ──────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def border_bottom(color=NAVY_HEX, weight="medium"):
    side = Side(style=weight, color=color)
    return Border(bottom=side)

def border_full(color="CCCCCC"):
    side = Side(style="thin", color=color)
    return Border(top=side, bottom=side, left=side, right=side)

def center():  return Alignment(horizontal="center", vertical="center")
def right():   return Alignment(horizontal="right",  vertical="center")
def left():    return Alignment(horizontal="left",   vertical="center")

def header_cell(ws, row, col, value, bg=NAVY_HEX):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=True, color=WHITE_HEX, size=10)
    c.alignment = center()
    c.border = border_full(WHITE_HEX)
    return c

def subheader_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(LIGHT_HEX)
    c.font = font(bold=True, color=NAVY_HEX, size=10)
    c.alignment = left()
    return c

def data_cell(ws, row, col, value, fmt=None, bg=WHITE_HEX, bold=False, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=color)
    c.alignment = right() if isinstance(value, (int, float)) else left()
    c.fill = fill(bg)
    if fmt:
        c.number_format = fmt
    c.border = border_full("DDDDDD")
    return c

# ── Build workbook ──────────────────────────────────────────────
def build_excel(output_path):
    t   = build_mortality_table()
    port= simulate_portfolio()
    tar = tarifier_portefeuille(port, t)
    sin = simulate_sinistres(tar, t, seed=12)
    cr_r= compte_resultat(sin)
    cr_t= compte_resultat_theorique(tar)
    mc  = monte_carlo_sp(tar, n_simulations=200, seed_base=0, table_mortalite=t)
    rv  = tarif_renouvellement(cr_r)
    
    wb = openpyxl.Workbook()
    
    # ════════════════════════════════════════════════
    # FEUILLE 1 : COMPTE DE RÉSULTAT
    # ════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Compte de résultat"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 38
    for col in ["B","C","D","E"]:
        ws1.column_dimensions[col].width = 16
    ws1.row_dimensions[1].height = 8
    
    # Titre
    ws1.merge_cells("A2:E2")
    t_cell = ws1.cell(row=2, column=1,
        value="COMPTE DE RÉSULTAT TECHNIQUE — RÉGIME DE PRÉVOYANCE COLLECTIVE")
    t_cell.fill = fill(NAVY_HEX)
    t_cell.font = Font(name="Calibri", bold=True, color=WHITE_HEX, size=13)
    t_cell.alignment = center()
    ws1.row_dimensions[2].height = 28
    
    ws1.merge_cells("A3:E3")
    s_cell = ws1.cell(row=3, column=1,
        value="Entreprise XYZ | 500 salariés | Exercice 2026")
    s_cell.fill = fill(BLUE_HEX)
    s_cell.font = Font(name="Calibri", bold=False, color=WHITE_HEX, size=10, italic=True)
    s_cell.alignment = center()
    ws1.row_dimensions[3].height = 18
    
    ws1.row_dimensions[4].height = 8
    
    # En-têtes colonnes
    headers = ["Poste", "Décès", "IJ", "Invalidité", "TOTAL"]
    for i, h in enumerate(headers, 1):
        header_cell(ws1, 5, i, h)
    ws1.row_dimensions[5].height = 20
    
    garanties = ["Décès", "IJ", "Invalidité"]
    
    # Section PRODUITS
    row = 6
    ws1.merge_cells(f"A{row}:E{row}")
    c = ws1.cell(row=row, column=1, value="PRODUITS")
    c.fill = fill(LIGHT_HEX)
    c.font = font(bold=True, color=NAVY_HEX, size=10)
    c.alignment = left()
    ws1.row_dimensions[row].height = 18
    
    row = 7
    ws1.cell(row=row, column=1, value="  Primes commerciales acquises (€)")
    ws1.cell(row=row, column=1).font = font(color="000000")
    ws1.cell(row=row, column=1).alignment = left()
    for j, g in enumerate(garanties, 2):
        data_cell(ws1, row, j, cr_t[g]['primes_commerciales'], "#,##0", GREY_HEX)
    data_cell(ws1, row, 5, cr_t['Total']['primes_commerciales'], "#,##0", GREY_HEX, bold=True, color=NAVY_HEX)
    
    row = 8
    ws1.cell(row=row, column=1, value="  dont primes pures (risque pur)")
    ws1.cell(row=row, column=1).font = font(italic=True, color="666666")
    ws1.cell(row=row, column=1).alignment = left()
    for j, g in enumerate(garanties, 2):
        data_cell(ws1, row, j, cr_t[g]['primes_pures'], "#,##0", WHITE_HEX)
    data_cell(ws1, row, 5, cr_t['Total']['primes_pures'], "#,##0", WHITE_HEX, bold=True)
    
    row = 9
    ws1.cell(row=row, column=1, value="  dont chargements de gestion")
    ws1.cell(row=row, column=1).font = font(italic=True, color="666666")
    ws1.cell(row=row, column=1).alignment = left()
    for j, g in enumerate(garanties, 2):
        data_cell(ws1, row, j, cr_t[g]['chargements'], "#,##0", WHITE_HEX)
    data_cell(ws1, row, 5, cr_t['Total']['chargements'], "#,##0", WHITE_HEX, bold=True)
    
    # Section CHARGES
    row = 10
    ws1.merge_cells(f"A{row}:E{row}")
    c = ws1.cell(row=row, column=1, value="CHARGES")
    c.fill = fill(LIGHT_HEX)
    c.font = font(bold=True, color=NAVY_HEX, size=10)
    c.alignment = left()
    ws1.row_dimensions[row].height = 18
    
    row = 11
    ws1.cell(row=row, column=1, value="  Sinistres réglés — exercice réalisé (€)")
    ws1.cell(row=row, column=1).font = font(color="000000")
    ws1.cell(row=row, column=1).alignment = left()
    for j, g in enumerate(garanties, 2):
        data_cell(ws1, row, j, cr_r[g]['sinistres'], "#,##0", GREY_HEX)
    data_cell(ws1, row, 5, cr_r['Total']['sinistres'], "#,##0", GREY_HEX, bold=True, color=ORANGE_HEX)
    
    row = 12
    ws1.cell(row=row, column=1, value="  Sinistres attendus — théoriques (€)")
    ws1.cell(row=row, column=1).font = font(italic=True, color="666666")
    ws1.cell(row=row, column=1).alignment = left()
    for j, g in enumerate(garanties, 2):
        data_cell(ws1, row, j, cr_t[g]['sinistres'], "#,##0", WHITE_HEX)
    data_cell(ws1, row, 5, cr_t['Total']['sinistres'], "#,##0", WHITE_HEX, bold=True)
    
    # Séparateur
    row = 13
    ws1.row_dimensions[row].height = 6
    
    # Résultats
    row = 14
    ws1.merge_cells(f"A{row}:E{row}")
    c = ws1.cell(row=row, column=1, value="RÉSULTATS")
    c.fill = fill(LIGHT_HEX)
    c.font = font(bold=True, color=NAVY_HEX, size=10)
    c.alignment = left()
    ws1.row_dimensions[row].height = 18
    
    row = 15
    ws1.cell(row=row, column=1, value="  Ratio S/P réalisé")
    ws1.cell(row=row, column=1).font = font(bold=True, color=NAVY_HEX)
    ws1.cell(row=row, column=1).alignment = left()
    for j, g in enumerate(garanties, 2):
        v = cr_r[g]['ratio_sp']
        c = data_cell(ws1, row, j, v, "0.0%", GREY_HEX)
        c.font = font(bold=True, color=(ORANGE_HEX if v > 0.90 else GREEN_HEX))
    v_tot = cr_r['Total']['ratio_sp']
    c = data_cell(ws1, row, 5, v_tot, "0.0%", GREY_HEX, bold=True)
    c.font = font(bold=True, size=11, color=(ORANGE_HEX if v_tot > 0.90 else GREEN_HEX))
    
    row = 16
    ws1.cell(row=row, column=1, value="  Ratio S/P théorique (attendu)")
    ws1.cell(row=row, column=1).font = font(italic=True, color="666666")
    ws1.cell(row=row, column=1).alignment = left()
    for j, g in enumerate(garanties, 2):
        data_cell(ws1, row, j, cr_t[g]['ratio_sp'], "0.0%", WHITE_HEX)
    data_cell(ws1, row, 5, cr_t['Total']['ratio_sp'], "0.0%", WHITE_HEX, bold=True)
    
    row = 17
    ws1.cell(row=row, column=1, value="  Résultat technique réalisé (€)")
    ws1.cell(row=row, column=1).font = font(bold=True)
    ws1.cell(row=row, column=1).alignment = left()
    for j, g in enumerate(garanties, 2):
        v = cr_r[g]['resultat']
        c = data_cell(ws1, row, j, v, "+#,##0;-#,##0;0", GREY_HEX, bold=True)
        c.font = font(bold=True, color=(GREEN_HEX if v >= 0 else ORANGE_HEX))
    v_tot = cr_r['Total']['resultat']
    c = data_cell(ws1, row, 5, v_tot, "+#,##0;-#,##0;0", GREY_HEX, bold=True)
    c.font = font(bold=True, size=11, color=(GREEN_HEX if v_tot >= 0 else ORANGE_HEX))
    
    # Bloc synthèse
    row = 19
    ws1.merge_cells(f"A{row}:E{row}")
    c = ws1.cell(row=row, column=1, value="INDICATEURS CLÉS DU RÉGIME")
    c.fill = fill(NAVY_HEX); c.font = font(bold=True, color=WHITE_HEX, size=10)
    c.alignment = center(); ws1.row_dimensions[row].height = 18
    
    kpis = [
        ("Effectif assuré", f"{len(tar):,} salariés"),
        ("Masse salariale annuelle", f"{tar['salaire_annuel'].sum():>12,.0f} €"),
        ("Taux de cotisation global", f"{tar['pc_totale'].sum()/tar['salaire_annuel'].sum()*100:.2f} % masse salariale"),
        ("Cotisation moyenne / salarié", f"{tar['pc_totale'].mean():,.0f} €/an"),
        ("Âge moyen du portefeuille", f"{tar['age'].mean():.1f} ans"),
        ("", ""),
        ("Renouvellement tarifaire recommandé", rv['interpretation']),
    ]
    for i, (label, val) in enumerate(kpis):
        r = row + 1 + i
        c = ws1.cell(row=r, column=1, value=label)
        c.font = font(bold=(label != ""))
        c.alignment = left()
        c.fill = fill(GREY_HEX if i % 2 == 0 else WHITE_HEX)
        v = ws1.cell(row=r, column=2, value=val)
        v.font = font(bold=True, color=NAVY_HEX)
        v.alignment = left()
        v.fill = fill(GREY_HEX if i % 2 == 0 else WHITE_HEX)
        ws1.merge_cells(f"B{r}:E{r}")
    
    # ════════════════════════════════════════════════
    # FEUILLE 2 : TARIFICATION PAR SALARIÉ
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet("Tarification")
    ws2.sheet_view.showGridLines = False
    
    cols_tar = {
        'id': ('ID', 5), 'age': ('Âge', 7), 'sexe': ('Sexe', 7),
        'salaire_mensuel': ('Salaire mensuel (€)', 16),
        'pp_deces':    ('PP Décès (€)',    14), 'pp_ij': ('PP IJ (€)', 12),
        'pp_invalidite': ('PP Inv. (€)',   13), 'pp_totale': ('PP Total (€)', 14),
        'pc_totale': ('PC Total (€)', 14), 'taux_cotisation_pct': ('Taux cot. (%)', 13),
    }
    
    # Title
    ws2.merge_cells(f"A1:{get_column_letter(len(cols_tar))}1")
    t2 = ws2.cell(row=1, column=1, value="DÉTAIL DE TARIFICATION PAR SALARIÉ — 500 premiers assurés")
    t2.fill = fill(NAVY_HEX); t2.font = font(bold=True, color=WHITE_HEX, size=12)
    t2.alignment = center(); ws2.row_dimensions[1].height = 24
    
    for i, (col_key, (col_label, col_width)) in enumerate(cols_tar.items(), 1):
        header_cell(ws2, 2, i, col_label)
        ws2.column_dimensions[get_column_letter(i)].width = col_width
    
    for row_idx, (_, row_data) in enumerate(tar.iterrows(), 3):
        bg = GREY_HEX if row_idx % 2 == 0 else WHITE_HEX
        for col_idx, col_key in enumerate(cols_tar.keys(), 1):
            val = row_data[col_key] if col_key != 'id' else row_data.name + 1
            fmt = None
            if col_key in ['salaire_mensuel', 'pp_deces', 'pp_ij', 'pp_invalidite', 'pp_totale', 'pc_totale']:
                fmt = "#,##0"
            elif col_key == 'taux_cotisation_pct':
                fmt = "0.00"
            data_cell(ws2, row_idx, col_idx, val, fmt, bg)
    
    # Ligne totaux
    n_data = len(tar) + 3
    ws2.row_dimensions[n_data].height = 20
    ws2.cell(row=n_data, column=1, value="TOTAL / MOYENNE")
    ws2.cell(row=n_data, column=1).font = font(bold=True, color=WHITE_HEX)
    ws2.cell(row=n_data, column=1).fill = fill(NAVY_HEX)
    
    totals = {
        'salaire_mensuel': tar['salaire_mensuel'].mean(),
        'pp_deces': tar['pp_deces'].sum(), 'pp_ij': tar['pp_ij'].sum(),
        'pp_invalidite': tar['pp_invalidite'].sum(), 'pp_totale': tar['pp_totale'].sum(),
        'pc_totale': tar['pc_totale'].sum(), 'taux_cotisation_pct': tar['taux_cotisation_pct'].mean(),
    }
    for col_idx, col_key in enumerate(cols_tar.keys(), 1):
        if col_key in totals:
            fmt = "#,##0" if col_key != 'taux_cotisation_pct' else "0.00"
            c = data_cell(ws2, n_data, col_idx, totals[col_key], fmt, NAVY_HEX, bold=True, color=WHITE_HEX)
    
    # ════════════════════════════════════════════════
    # FEUILLE 3 : TABLES ACTUARIELLES
    # ════════════════════════════════════════════════
    ws3 = wb.create_sheet("Tables actuarielles")
    ws3.sheet_view.showGridLines = False
    
    ws3.merge_cells("A1:F1")
    t3 = ws3.cell(row=1, column=1, value="TABLES ACTUARIELLES — TD 88-90 & MORBIDITÉ")
    t3.fill = fill(NAVY_HEX); t3.font = font(bold=True, color=WHITE_HEX, size=12)
    t3.alignment = center(); ws3.row_dimensions[1].height = 24
    
    mort_table = build_mortality_table(22, 65)
    mort_headers = ["Âge", "qx Hommes (‰)", "qx Femmes (‰)", "Ratio H/F",
                    "p_arrêt (%)", "Durée moy IJ (j)", "p_invalidité (‰)", "ä_{65-x}"]
    col_widths = [8, 16, 16, 12, 14, 18, 18, 14]
    
    for i, (h, w) in enumerate(zip(mort_headers, col_widths), 1):
        header_cell(ws3, 2, i, h)
        ws3.column_dimensions[get_column_letter(i)].width = w
    
    from model_prevoyance import get_morbidite, annuite_certaine
    for row_i, age in enumerate(range(22, 66), 3):
        bg = GREY_HEX if row_i % 2 == 0 else WHITE_HEX
        qxh = mort_table.loc[age, 'qx_homme']
        qxf = mort_table.loc[age, 'qx_femme']
        pa, dij, pinv = get_morbidite(age)
        ann = annuite_certaine(65 - age)
        row_vals = [age, round(qxh*1000,3), round(qxf*1000,3), round(qxh/qxf,2),
                    round(pa*100,1), dij, round(pinv*1000,2), round(ann,3)]
        fmts = [None,"0.000","0.000","0.00","0.0",None,"0.000","0.000"]
        for ci, (v, f) in enumerate(zip(row_vals, fmts), 1):
            data_cell(ws3, row_i, ci, v, f, bg)
    
    # ════════════════════════════════════════════════
    # FEUILLE 4 : MONTE CARLO
    # ════════════════════════════════════════════════
    ws4 = wb.create_sheet("Monte Carlo SP")
    ws4.sheet_view.showGridLines = False
    
    ws4.merge_cells("A1:E1")
    t4 = ws4.cell(row=1, column=1,
        value="DISTRIBUTION DU RATIO S/P — 200 SIMULATIONS (Monte Carlo)")
    t4.fill = fill(NAVY_HEX); t4.font = font(bold=True, color=WHITE_HEX, size=12)
    t4.alignment = center(); ws4.row_dimensions[1].height = 24
    
    # Statistics
    sp_vals = mc['sp_total']
    stats = [
        ("Nombre de simulations", len(mc)),
        ("S/P moyen (= S/P théorique attendu)", f"{sp_vals.mean():.1%}"),
        ("S/P médian", f"{sp_vals.median():.1%}"),
        ("Percentile 5% (année favorable)", f"{sp_vals.quantile(0.05):.1%}"),
        ("Percentile 95% (année défavorable)", f"{sp_vals.quantile(0.95):.1%}"),
        ("% années avec S/P > 90%", f"{(sp_vals > 0.90).mean():.1%}"),
        ("% années déficitaires (S/P > 100%)", f"{(sp_vals > 1.00).mean():.1%}"),
        ("Résultat moyen (€)", f"{mc['resultat'].mean():,.0f} €"),
        ("Résultat min (€)", f"{mc['resultat'].min():,.0f} €"),
        ("Résultat max (€)", f"{mc['resultat'].max():,.0f} €"),
    ]
    
    ws4.row_dimensions[2].height = 6
    for i, (label, val) in enumerate(stats, 3):
        bg = GREY_HEX if i % 2 != 0 else WHITE_HEX
        c = ws4.cell(row=i, column=1, value=label)
        c.font = font(bold=True); c.alignment = left(); c.fill = fill(bg)
        ws4.column_dimensions["A"].width = 45
        v = ws4.cell(row=i, column=2, value=val)
        v.font = font(bold=True, color=NAVY_HEX); v.alignment = left(); v.fill = fill(bg)
        ws4.column_dimensions["B"].width = 22
    
    # Data table
    row = len(stats) + 5
    ws4.merge_cells(f"A{row}:E{row}")
    c = ws4.cell(row=row, column=1, value="Détail des 200 simulations")
    c.fill = fill(BLUE_HEX); c.font = font(bold=True, color=WHITE_HEX)
    c.alignment = center(); ws4.row_dimensions[row].height = 18
    
    mc_headers = ["Sim #", "S/P Total", "S/P Décès", "S/P IJ", "S/P Invalidité", "Résultat (€)"]
    mc_cols = ["sim", "sp_total", "sp_deces", "sp_ij", "sp_invalidite", "resultat"]
    col_w2 = [8, 12, 12, 12, 14, 15]
    for i, (h, w) in enumerate(zip(mc_headers, col_w2), 1):
        header_cell(ws4, row+1, i, h, BLUE_HEX)
        ws4.column_dimensions[get_column_letter(i)].width = w
    
    for ri, (_, mc_row) in enumerate(mc.iterrows(), row+2):
        bg = GREY_HEX if ri % 2 == 0 else WHITE_HEX
        sp_val = mc_row['sp_total']
        for ci, col in enumerate(mc_cols, 1):
            v = mc_row[col]
            if col == 'sim': 
                fmt = None; val = int(v)+1
            elif col == 'resultat': 
                fmt = "+#,##0;-#,##0;0"; val = v
            else: 
                fmt = "0.0%"; val = v
            c = data_cell(ws4, ri, ci, val, fmt, bg)
            if col == 'sp_total':
                c.fill = fill("FFD9D9" if sp_val > 1.0 else ("FFF2CC" if sp_val > 0.90 else "E2EFDA"))
                c.font = font(bold=True, color=(ORANGE_HEX if sp_val > 1.0 else ("996600" if sp_val > 0.90 else GREEN_HEX)))
    
    wb.save(output_path)
    print(f"✅ Excel généré : {output_path}")
    print(f"   4 feuilles : Compte de résultat | Tarification | Tables actuarielles | Monte Carlo S/P")

if __name__ == "__main__":
    out = "/mnt/user-data/outputs/compte_resultat_prevoyance.xlsx"
    build_excel(out)
